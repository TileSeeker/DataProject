# -*- coding: utf-8 -*-
"""
Created on Tue Apr 13 16:09:00 2021

@author: emism
"""
from strømpriser import strømpris
from Signals import Signal
import time
import openpyxl
from datetime import datetime
from datetime import date

from forbruksModell import userPowerConsumption

#CoT-sikringsskap
token = "eyJhbGciOiJIUzI1NiJ9.eyJqdGkiOiI1ODMyIn0.1sZoVBXJzmhJ71UiSxFxPcZ-3FufHWOO1UokVe0IBwA"
key_rom1 = "24060"
key_rom2 = "9511"
key_rom3 = "12024"
key_rom4 = "2953"
key_rom5 = "17975"
key_rom6 = "4649"

#CoT-brukere + key
token1 = "eyJhbGciOiJIUzI1NiJ9.eyJqdGkiOiI1NTk1In0.9bD-g6Gi40yjEiEYGOY1eoWl0wEuAZZN67yzS5gYOQs"
token2 = "eyJhbGciOiJIUzI1NiJ9.eyJqdGkiOiI1NTk2In0.zADZlTeWjpJpbEmG_d1mcw07mDtT9ZJ30sMDtU1ex80"
token3 = "eyJhbGciOiJIUzI1NiJ9.eyJqdGkiOiI1NTk3In0.es9iHyTEfrYM3ksN0QWtiULhRlQEcwatXWHFc5_fscc"
token4 = "eyJhbGciOiJIUzI1NiJ9.eyJqdGkiOiI1NTk4In0.WP5pJqMaPL8AwEEii2TMFys9kQUabpl2iztyxBRdLuc"
token5 = "eyJhbGciOiJIUzI1NiJ9.eyJqdGkiOiI1NTk5In0.fbZfuVuDshnMdUMW6EXrB6fSYhtdq0l2-j92h6AtlbM"
token6 = "eyJhbGciOiJIUzI1NiJ9.eyJqdGkiOiI1NjAwIn0.nGZXNRU34wVFtzex9tS-0gVky_ppn3Gjmj_riA4oLZY"
key_1 = "1069"
key_2 = "4641"
key_3 = "16199"
key_4 = "20880"
key_5 = "19096"
key_6 = "15853"

#excel headers
headers = ("Day/hour", "Wh_1", "Wh_2", "Wh_3", "Wh_4", "Wh_5", "Wh_6", 
            "ore_1", "ore_2", "ore_3", "ore_4", "ore_5", "ore_6", 
            "t_kWh_1", "t_kWh_2", "t_kWh_3", "t_kWh_4", "t_kWh_5", "t_kWh_6",
            "t_NOK_1","t_NOK_2","t_NOK_3","t_NOK_4","t_NOK_5","t_NOK_6",)

#1 gang i timen
def kWh_to_NOK_upload():
    todays_date = date.today()
    current_date = str(todays_date.year) + "-" + str(todays_date.month)
    
    wb = openpyxl.load_workbook("strømbruk.xlsx", data_only=True)
    sh = wb[current_date]

    NOK_1 =sh.cell(2,20).value
    NOK_2 =sh.cell(2,21).value
    NOK_3 =sh.cell(2,22).value
    NOK_4 =sh.cell(2,23).value
    NOK_5 =sh.cell(2,24).value
    NOK_6 =sh.cell(2,25).value
    
    Signal(key_1, token1).write(NOK_1)
    Signal(key_2, token2).write(NOK_2)
    Signal(key_3, token3).write(NOK_3)
    Signal(key_4, token4).write(NOK_4)
    Signal(key_5, token5).write(NOK_5)
    Signal(key_6, token6).write(NOK_6)
    wb.save("strømbruk.xlsx")

#1 gang i timen
def get_store_data():
    #get data from cot
    signal1 = Signal(key_rom1, token).get()
    signal2 = Signal(key_rom2, token).get()
    signal3 = Signal(key_rom3, token).get()
    signal4 = Signal(key_rom4, token).get()
    signal5 = Signal(key_rom5, token).get()
    signal6 = Signal(key_rom6, token).get()
    time_now = datetime.now().strftime("%d %H:%M:%S")
    
    
    #dagens_strømpris = strømpris("D:\.drivere\geckodriver.exe")
    #dagens_strømpris = dagens_strømpris.get()
    dagens_strømpris = userPowerConsumption().recentPowerPrice()
    
    #dagens_strømpris = 22.5 # ore/kWh
    #strømpris*wh*10
    ore1 = (signal1*0.001)*dagens_strømpris
    ore2 = (signal2*0.001)*dagens_strømpris
    ore3 = (signal3*0.001)*dagens_strømpris
    ore4 = (signal4*0.001)*dagens_strømpris
    ore5 = (signal5*0.001)*dagens_strømpris
    ore6 = (signal6*0.001)*dagens_strømpris

    strømlist = (time_now, signal1, signal2, signal3, signal4, signal5, signal6, ore1, ore2, ore3, ore4, ore5, ore6)

    #store data in excel
    wb = openpyxl.load_workbook("strømbruk.xlsx")
    todays_date = date.today()
    current_date = str(todays_date.year) + "-" + str(todays_date.month)
    sh = wb[current_date]
    sh.append(strømlist)
    wb.save("strømbruk.xlsx")

#creating datasheet for 1 month # 1 gang i måneden
def create_datasheet():
    wb = openpyxl.load_workbook("strømbruk.xlsx")
    todays_date = date.today()
    current_date = str(todays_date.year) + "-" + str(todays_date.month)
    
    if str(current_date) in wb.sheetnames:
        print("Sheet already exsist")
        
    else:
        wb.create_sheet(current_date)
        sh = wb[current_date]
        sh.append(headers)
    
        #sum for each person in kWh
        sh["Z1"] = "0.001"
        sh["N2"] = "=SUM(B2:B747)*Z1"
        sh["O2"] = "=SUM(C2:C747)*Z1"
        sh["P2"] = "=SUM(D2:D747)*Z1"
        sh["Q2"] = "=SUM(E2:E747)*Z1"
        sh["R2"] = "=SUM(F2:F747)*Z1"
        sh["S2"] = "=SUM(G2:G747)*Z1"
        
        #sum for each person in NOK
        sh["AA1"] = "0.01"
        sh["T2"] = "=SUM(H2:H747)*AA1"
        sh["U2"] = "=SUM(I2:I747)*AA1"
        sh["V2"] = "=SUM(J2:J747)*AA1"
        sh["W2"] = "=SUM(K2:K747)*AA1"
        sh["X2"] = "=SUM(L2:L747)*AA1"
        sh["Y2"] = "=SUM(M2:M747)*AA1"
        print("New Sheet Created")


    wb.save("strømbruk.xlsx")
    
if __name__ == "__main__":
    create_datasheet()
    get_store_data()
